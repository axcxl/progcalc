import argparse
from guizero import App, TextBox, Text, Combo, Waffle, Box, ListBox, Window
import openpyxl


class ProgCalc:
    def __init__(self, input_excel_db):
        self.min_size = 32
        self.no_bits = 8
        self.value = 0
        self.bit_map = {}
        self.bit_descr = {}
        self.endianess = "big"

        self.app = App(layout="grid", height=350, width=550)
        self.top_box = Box(self.app, layout = "grid", grid = [0, 0])
        self.bottom_box = Box(self.app, layout = "grid", grid = [0, 1])
        self.right_box = Box(self.app, grid = [1, 0, 1, 2])

        self.window = Window(self.app, width=250, height=150, visible=False)
        self.description = Text(self.window)

        # Create the text field to enter data
        self.input = TextBox(self.top_box, width=25, grid=[0,0,3,1], command=self.process_input)
        self.input.bg = "white"
        self.input.text_size = 16

        # Display the hex value
        self.out_hex = Text(self.top_box, grid=[2,1], text="0x<waiting for valid input>")
        # Display the min number of bits selector
        self.in_minsize = Combo(self.top_box, grid=[0, 1], options=["32"], command=self.process_minsize, selected="32")
        # Endianess selector
        self.in_endianess = Combo(self.top_box, grid=[1, 1], options=["little", "big"], command=self.process_endianess,
                                  selected=self.endianess)

        # Display the binary value
        self.out_bin = Text(self.top_box, grid=[0, 2, 2, 1], size=17, text="0b<waiting for valid input>")
        # Display little number after the binary value
        self.out_num = Text(self.top_box, grid=[0, 3, 2, 1], size=7, text="")


        # Prepare the waffle list
        self.waffle_list = []
        self.box_list = []
        self.text_list = []

        self.append_wb(4)
        self.append_tx(self.no_bits)

        # Read the worksheets in the input dictionary and create the list
        self.in_excel = openpyxl.load_workbook(filename = input_excel_db, read_only=True)
        self.in_regs = self.in_excel.sheetnames
        self.in_regs.insert(0, "OFF")

        # Display the list of registers
        self.in_reglist = ListBox(self.right_box, items = self.in_regs, command=self.process_reglist)

        self.input.focus()
        self.app.display()

    def append_wb(self, number):
        last_waffle = len(self.waffle_list)
        for i in range(0, 2* number, 2):
            w = Waffle(self.bottom_box, height=self.no_bits, width=1, grid=[last_waffle + i, 0])
            w.hide()
            w.when_clicked = self.process_waffle
            self.waffle_list.append(w)

            b = Box(self.bottom_box, grid=[i + 1, 0])
            b.hide()
            self.box_list.append(b)

    def append_tx(self, number):
        for i in range(0, len(self.box_list)):
            try:
                tmp = self.text_list[i]
            except:
                tmp = []

            for j in range(0, self.no_bits):
                tx = Text(self.box_list[i])
                tx.hide()
                tx.size = 14 # perfect size for all variants
                tx.when_mouse_enters = self.show_description
                tx.when_mouse_leaves = self.hide_description
                tmp.append(tx)
            self.text_list.append(tmp)

    def show_description(self, event_data):
        name = event_data.widget.value
        for elem in self.bit_map:
            if self.bit_map[elem] == name:
                self.description.value = self.bit_descr[elem]
                break
        self.window.show()

    def hide_description(self, event_data):
        self.window.hide()

    def process_input(self, inp):
        try:
            self.value = int(self.input.value, 0)
            self.refresh_all()
        except ValueError:
            return

    def process_endianess(self, inp):
        self.endianess = inp
        self.refresh_all()

    def process_minsize(self, opt):
        self.min_size = int(opt)
        self.refresh_all()

    def process_waffle(self, event_data):
        # Not sure if this is the best way, but it works
        waffleno = int(event_data.y / (event_data.widget.pixel_size + event_data.widget.pad))
        waffleno = event_data.widget.height - (waffleno + 1)

        # Get WafflePixel based on calculation above - reversed to take into account multiple waffles
        idx = 0
        for w in reversed(self.waffle_list):
            # Skip over invisible waffles
            if w.visible == False:
                continue
            if event_data.widget == w:
                break
            idx += 1

        # See what bit needs to be changed
        bit = self.no_bits * idx + waffleno
        if self.value & (1 << bit):
            self.value &= ~(1 << bit)
        else:
            self.value |= (1 << bit)

        # Refresh display
        self.refresh_all()

    def process_reglist(self, selected):
        if selected == "OFF":
            for b in self.box_list:
                b.hide()
            self.bit_map = {}
            self.refresh_all()
            return

        tmp = self.in_excel[selected]
        self.bit_map = {}
        for row in tmp.iter_rows():
            try:
                bits = row[0].value
                if bits == None:
                    break
            except IndexError:
                break
            name = row[1].value
            descr = row[2].value

            try:
                bit = int(bits)
                self.bit_map[bit] = name
                self.bit_descr[bit] = descr
            except ValueError:
                if bits == "Bits":
                    continue
                else:
                    bits = str.replace(bits, "–", "-")
                    interval = bits.split("-")
                    for i in range(int(interval[0]), int(interval[1]) + 1):
                        self.bit_map[i] = name
                        self.bit_descr[i] = descr
        self.refresh_all()

    def refresh_all(self):
        if "0x" in self.input.value:
            self.input.value = hex(self.value)
        else:
            self.input.value = str(self.value)

        # TODO: improve?
        self.display_hex()
        self.display_bin()
        self.display_waffle()

    def display_hex(self):
        self.out_hex.value = hex(self.value)

    def display_bin(self):
        # Get binary without 0b
        outstring = bin(self.value)[2:]
        n = len(outstring)

        # Add zeros at beginning if needed
        while n % self.min_size != 0:
            outstring = "0" + outstring
            n += 1

        # Split in groups of no_bits
        tmp = ""
        for idx in range(0, len(outstring)):
            tmp += outstring[idx]
            if (idx + 1) % self.no_bits == 0:
                tmp += " "
        outstring = tmp[:-1]

        # Handles endianess
        if self.endianess == "little":
            tmp = ""
            for elem in reversed(outstring.split(" ")):
                tmp += elem + " "
            outstring = tmp

        self.out_num.value = "0" + (" " * 140) + "31"
        self.out_bin.value = outstring

    def display_waffle(self):
        # Display new ones, based on the binary representation of the number
        idx = 0
        x_coord = 0

        # Dinamically add waffles if needed
        no_waffles = len(self.out_bin.value.split(" "))
        if no_waffles > len(self.waffle_list):
            self.append_wb(no_waffles - len(self.waffle_list))

        for elem in self.out_bin.value.split(" "):
            # Somehow it gets a empty element, fix this
            if elem == "":
                break

            # Get the waffle
            w = self.waffle_list[x_coord]
            w.show() # Display it
            if len(self.bit_map) != 0:
                b = self.box_list[x_coord]
                b.show()
                txlist = self.text_list[x_coord]
            # Set the individual pixel
            for i in range(0, len(elem)):
                if len(self.bit_map) != 0:
                    txlist[i].show()
                    txlist[i].value = str(idx) + ": " + self.bit_map[idx] + " " * 5
                    idx += 1
                wp = w.pixel(0, i)
                if elem[i] == "1":
                    wp.color = "blue"
                else:
                    wp.color = "white"

            x_coord += 1

        for i in range(x_coord, len(self.waffle_list)):
            self.waffle_list[i].hide()
            self.box_list[i].hide()

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Special calculator that interprets registers")
    parser.add_argument('excel_db', help='Path to an excel database file')

    args = parser.parse_args()

    pc = ProgCalc(args.excel_db)
